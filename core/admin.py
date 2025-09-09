from django.contrib import admin
from django.contrib.admin import DateFieldListFilter
from django.http import HttpResponse, HttpResponseRedirect
from .models import Policial, Producao, Formacao, Participacao
from .menu import RankingPlaceholder, ExportacaoPlaceholder
import os
import csv
from datetime import datetime
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# 👮 Admin de Policial
@admin.register(Policial)
class PolicialAdmin(admin.ModelAdmin):
    list_display = ('nome_guerra', 'graduacao', 're', 'pelotao')
    search_fields = ('nome_guerra', 're')

    def has_add_permission(self, request):
        if request.path == "/admin/":
            return False
        return super().has_add_permission(request)


# 📋 Admin de Produção
@admin.register(Producao)
class ProducaoAdmin(admin.ModelAdmin):
    autocomplete_fields = ['policial']
    list_display = ('policial', 'data', 'pontuacao_formatada')
    list_filter = (('data', DateFieldListFilter), 'policial__pelotao')
    search_fields = ('policial__nome_guerra', 'policial__re')
    actions = ['exportar_excel', 'exportar_csv']

    def has_add_permission(self, request):
        if request.path == "/admin/":
            return False
        return super().has_add_permission(request)

    # ⚠️ ACTION: precisa receber (self, request, queryset)
    def exportar_excel(self, request, queryset):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Produção'

        headers = [
            'Data', 'Policial', 'RE', 'Pelotão',
            'Pessoas', 'Pessoas AISP', 'Carros', 'Carros AISP',
            'Motos', 'Motos AISP', 'Ocorrências', 'Flagrantes',
            'Flagrantes AISP', 'Autuações', 'Raia', 'Procurado',
            'Carros Apreendidos', 'Motos Apreendidas', 'Flagrantes Outros',
            'Armas', 'Escolas', 'Pontuação'
        ]
        ws.append(headers)

        for p in queryset:
            # garante número/decimal como float e datas formatadas
            pont = float(p.pontuacao) if isinstance(p.pontuacao, Decimal) else round(p.pontuacao, 2)
            ws.append([
                p.data,  # deixamos como date; abaixo seto number_format
                p.policial.nome_guerra,
                p.policial.re,
                p.policial.pelotao,
                p.pessoa,
                p.pessoas_aisp,
                p.carros,
                p.carros_aisp,
                p.motos,
                p.motos_aisp,
                p.qnt_ocorrencias,
                p.flagrantes,
                p.flagrantes_aisp,
                p.autuacoes,
                p.raia,
                p.procurado,
                p.carro_apreendido,
                p.moto_apreendida,
                p.flagrantes_outros,
                p.arma,
                p.escolas,
                round(pont, 2),
            ])

        # Formatação: cabeçalho centralizado e largura aproximada
        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 16
            ws[f'{col_letter}1'].alignment = Alignment(horizontal='center', vertical='center')

        # Formato de data na coluna A
        for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
            for cell in row:
                cell.number_format = 'yyyy-mm-dd'

        # Congela cabeçalho
        ws.freeze_panes = 'A2'

        # 🟩 Caminho histórico
        pasta_exportacao = r'C:\relatorios\producoes'
        os.makedirs(pasta_exportacao, exist_ok=True)
        nome_arquivo = f"producao_{datetime.now().strftime('%Y-%m-%d_%Hh%Mm')}.xlsx"
        caminho_historico = os.path.join(pasta_exportacao, nome_arquivo)
        wb.save(caminho_historico)

        # 🟨 Caminho fixo Power BI
        pasta_powerbi = r'C:\relatorios\powerbi'
        os.makedirs(pasta_powerbi, exist_ok=True)
        caminho_powerbi = os.path.join(pasta_powerbi, 'producao_powerbi.xlsx')
        wb.save(caminho_powerbi)

        # ⬇️ Retorna também como download
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'
        wb.save(response)
        return response

    # (opcional) descrição amigável no menu de ações
    exportar_excel.short_description = "Exportar Excel (salva histórico, Power BI e baixa)"

    def exportar_csv(self, request, queryset):
        import os
        from datetime import datetime
        from django.http import HttpResponse

        pasta_exportacao = r'C:\relatorios\producoes'
        pasta_powerbi = r'C:\relatorios\powerbi'
        os.makedirs(pasta_exportacao, exist_ok=True)
        os.makedirs(pasta_powerbi, exist_ok=True)

        nome_arquivo = f"producao_{datetime.now().strftime('%Y-%m-%d_%Hh%Mm')}.csv"
        caminho_historico = os.path.join(pasta_exportacao, nome_arquivo)
        caminho_powerbi = os.path.join(pasta_powerbi, 'producao_powerbi.csv')

        campos = [
            'data', 'nome_guerra', 're', 'pelotao',
            'pessoa', 'pessoas_aisp', 'carros', 'carros_aisp',
            'motos', 'motos_aisp', 'qnt_ocorrencias', 'flagrantes',
            'flagrantes_aisp', 'autuacoes', 'raia', 'procurado',
            'carro_apreendido', 'moto_apreendida', 'flagrantes_outros',
            'arma', 'escolas', 'pontuacao'
        ]

        def _linha(p):
            pont = float(p.pontuacao) if isinstance(p.pontuacao, Decimal) else round(p.pontuacao, 2)
            return [
                p.data.strftime('%Y-%m-%d'),
                p.policial.nome_guerra, p.policial.re, p.policial.pelotao,
                p.pessoa, p.pessoas_aisp, p.carros, p.carros_aisp,
                p.motos, p.motos_aisp, p.qnt_ocorrencias, p.flagrantes,
                p.flagrantes_aisp, p.autuacoes, p.raia, p.procurado,
                p.carro_apreendido, p.moto_apreendida, p.flagrantes_outros,
                p.arma, p.escolas, round(pont, 2)
            ]

        # 🟩 Arquivo histórico
        with open(caminho_historico, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(campos)
            for p in queryset:
                writer.writerow(_linha(p))

        # 🟨 Arquivo fixo Power BI
        with open(caminho_powerbi, 'w', newline='', encoding='utf-8') as f2:
            writer = csv.writer(f2)
            writer.writerow(campos)
            for p in queryset:
                writer.writerow(_linha(p))

        # ⬇️ Download imediato também
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'
        writer = csv.writer(response)
        writer.writerow(campos)
        for p in queryset:
            writer.writerow(_linha(p))
        return response

    def pontuacao_formatada(self, obj):
        return f"{obj.pontuacao:.2f}"
    # (tinha um typo aqui)
    pontuacao_formatada.short_description = "Pontuação"


@admin.register(EquipeProdutividade)
class EquipeProdutividadeAdmin(admin.ModelAdmin):
    list_display = ('data', 'pelotao', 'pontuacao', 'gerou_individual')
    filter_horizontal = ('policiais',)
    list_filter = ('pelotao', 'data', 'gerou_individual')
    search_fields = ('pelotao',)

    def save_model(self, request, obj, form, change):
        super().save_model(request, obj, form, change)

        if obj.gerou_individual:
            # ⚠️ Remove produções antigas para evitar duplicação
            Producao.objects.filter(data=obj.data, policial__in=obj.policiais.all()).delete()

            # 🚀 Gera novamente para cada policial
            for policial in obj.policiais.all():
                Producao.objects.create(
                    policial=policial,
                    data=obj.data,
                    pessoa=obj.pessoa,
                    carros=obj.carros,
                    motos=obj.motos,
                    pessoas_aisp=obj.pessoas_aisp,
                    carros_aisp=obj.carros_aisp,
                    motos_aisp=obj.motos_aisp,
                    qnt_ocorrencias=obj.qnt_ocorrencias,
                    flagrantes=obj.flagrantes,
                    flagrantes_aisp=obj.flagrantes_aisp,
                    autuacoes=obj.autuacoes,
                    raia=obj.raia,
                    procurado=obj.procurado,
                    carro_apreendido=obj.carro_apreendido,
                    moto_apreendida=obj.moto_apreendida,
                    flagrantes_outros=obj.flagrantes_outros,
                    arma=obj.arma,
                    escolas=obj.escolas,
                    observacao=f"(Equipe) {obj.observacao or ''}".strip(),
                )

            self.message_user(request, f"✔ Produção individual atualizada para {obj.policiais.count()} policiais.")

        else:
            self.message_user(request, "ℹ️ Produção individual **não** foi gerada. Marque 'gerou individual' para habilitar esse recurso.", level=messages.WARNING)